import os
import shutil
import pandas as pd
import openpyxl

def process_excel(excel_file, output_file_path):
    df = pd.read_excel(excel_file)
    row_index = 0

    # Create a new Excel workbook
    workbook = openpyxl.Workbook()

    while row_index < len(df):
        row = df.iloc[row_index]

        # Check for required columns
        if 'Source' in df.columns and 'Destination' in df.columns and 'OldWord' in df.columns and 'NewWord' in df.columns:
            source_folder = row['Source']
            dest_folder = row['Destination']
            old_word = row['OldWord']
            new_word = row['NewWord']

            # Check if destination folder already exists
            if not os.path.exists(dest_folder):
                os.makedirs(dest_folder)
                print(f"Created destination folder: {dest_folder}")
            else:
                print(f"Destination folder already exists: {dest_folder}. Skipping copying.")

            # Rename files within the destination folder (if it exists)
            if os.path.isdir(dest_folder):
                for filename in os.listdir(dest_folder):
                    if old_word in filename:
                        new_filename = filename.replace(old_word, new_word)
                        os.rename(os.path.join(dest_folder, filename), os.path.join(dest_folder, new_filename))
                        print(f"Renamed file {filename} to {new_filename} in {dest_folder}")

                # Create a new sheet for the destination folder
                sheet = workbook.create_sheet(os.path.basename(dest_folder))

                # Populate the sheet with file names and hyperlinks
                row_num = 1
                for file in os.listdir(dest_folder):
                    file_path = os.path.join(dest_folder, file)
                    if os.path.isfile(file_path):
                        sheet.cell(row=row_num, column=1, value=file)
                        sheet.cell(row=row_num, column=1).hyperlink = file_path
                        row_num += 1

        row_index += 1

    # Save the workbook
    workbook.save(output_file_path)

# Example usage
excel_file_path = r'c:\script\rmg\CP-Ren-UL.xlsx'
output_file_path = r"C:\Office\OneDrive - ATOS\Desktop\123_Prolinks.xlsx"

process_excel(excel_file_path, output_file_path)
